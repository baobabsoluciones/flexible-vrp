# Class to solve the problem with a basic mip.
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side

from flexible_vrp.core import Experiment, Solution
from flexible_vrp.solver.basic_mip_tools.create_model import create_model
from pyomo.environ import SolverFactory
from cornflow_client.constants import PYOMO_STOP_MAPPING
from pytups import TupList


class BasicMip(Experiment):
    def __init__(self, instance, solution=None):
        super().__init__(instance, solution)

    def prepare_model_data(self):
        # data is the dict that is filled in and returned
        # Calculating MILP parameters and sets
        # Generating the dictionary for Pyomo

        data = dict()

        # Adding the set for Warehouses
        # Retrieving origins and destination
        origins = set([c["location1"] for c in self.instance.data["trip_durations"]])
        destinations = set(
            [c["location2"] for c in self.instance.data["trip_durations"]]
        )
        # Getting a list from the union of destinations and origins
        warehouses = list(origins.union(destinations))
        # Adding the info to the output dict
        data["sWarehouses"] = {None: warehouses}

        # Adding the set for Commodities
        data["sCommodities"] = {
            None: [
                (c["origin"], c["destination"], c["quantity"], c["required"])
                for c in self.instance.data["commodities"]
            ]
        }
        # todo: estimate the number of stops for the current data. Remove from this method
        # self.instance.data["parameters"]["no_stops"] = 22
        # Adding the set for Stops
        data["sStops"] = {
            None: [s for s in range(int(self.instance.data["parameters"]["no_stops"]))]
        }
        data["sStopsButLast"] = {
            None: [
                s for s in range(int(self.instance.data["parameters"]["no_stops"] - 1))
            ]
        }
        # Adding the set for Vehicles
        data["sVehicles"] = {
            None: [
                v for v in range(int(self.instance.data["parameters"]["fleet_size"]))
            ]
        }

        data["sTripDuration"] = {
            None: [
                (v, s, s + 1) for v in range(int(self.instance.data["parameters"]["fleet_size"]))
                for s in range(int(self.instance.data["parameters"]["no_stops"] - 1))
            ]
        }
        # Adding the parameters
        data["pVehCAP"] = {None: self.instance.data["parameters"]["vehicle_capacity"]}
        data["pFleetSize"] = {None: self.instance.data["parameters"]["fleet_size"]}
        data["pLoadTime"] = {None: self.instance.data["parameters"]["load_pallet"]}
        data["pUnloadTime"] = {None: self.instance.data["parameters"]["unload_pallet"]}
        data["pReqTimeLimit"] = {
            None: self.instance.data["parameters"]["req_time_limit"]
        }
        data["pOptTimeLimit"] = {
            None: self.instance.data["parameters"]["opt_time_limit"]
        }

        # Adding the pCommodities parameter
        data["pTripDuration"] = {
            (x["location1"], x["location2"]): x["time"]
            for x in self.instance.data["trip_durations"]
        }

        # Adding the parameters BigM

        data["pBigM1"] = {None: max(data["pTripDuration"].values())}
        data["pBigM2"] = data["pOptTimeLimit"]
        data["pBigM3"] = {None: data["pOptTimeLimit"][None] - data["pReqTimeLimit"][None]}

        return {None: data}

    def solve(self, options):
        if not self.instance.to_dict():
            raise ValueError("The instance is empty")

        data = self.prepare_model_data()

        # Todo: replace this by the solve method
        model = create_model()

        model_instance = model.create_instance(data, report_timing=False)
        print("number constraints", model_instance.nconstraints())
        logfile = "./data/logfile.log"

        # Solve
        opt = self.set_solver(options)
        result = opt.solve(
            model_instance,
            tee=True,
            warmstart=False,
            logfile=logfile,
        )
        status = PYOMO_STOP_MAPPING[result.solver.termination_condition]
        data_json = self.get_solution(model_instance)
        self.solution = Solution({"data": data_json})

        # Printing the value of the OF
        obj = model_instance.obj_func()
        print("Status: {} Objective value: {}".format(status, obj))

        # # Prepare solution
        # if self.is_feasible(self.status): #todo: create method
        #     self.totals = self.get_total(model_result, result)
        #     model_solution = self.get_model_solution(model_result)
        #     self.solution = Solution(self.build_solution(model_solution))
        # else:
        #     self.solution = self.get_empty_solution()
        #     self.variables_to_excel(model_result)

        return 1  # dict(status=STATUS_TIME_LIMIT, status_sol=SOLUTION_STATUS_FEASIBLE)

    def set_solver(self, options):
        # Create the solver object and set the relevant options.
        #
        # param options: dict of options.
        # :return: the pyomo solver
        if "solver_name" in options:
            opt = SolverFactory(options["solver_name"])
        else:
            opt = SolverFactory("cbc")
        if "solver_config" in options:
            opt.options.update(options["solver_config"])
        return opt

    def get_solution(self, model_instance):

        data_dep_time = TupList([[  # df que recoge los minutos de salida de cada w
            model_instance.vDepartureTime[v, s].value]
            for v in model_instance.sVehicles
            for s in model_instance.sStops
            for w in model_instance.sWarehouses
            for c in model_instance.sCommodities
            if model_instance.vLoadQuantity[v, s, c].value + model_instance.vUnloadQuantity[v, s, c].value +
            model_instance.vQuantityAtArrival[v, s, c].value > 0 and model_instance.vAlpha[v, s, w].value == 1
        ]).to_dictlist(["T_salida"])

        t_utilizado = "{:.1f}".format(data_dep_time[-1]["T_salida"] / 60)  # vble que muestra las horas utilizadas

        data_arr_time = TupList([[  # df que muestra los minutos de entrada a cada w
            model_instance.vArrivalTime[v, s].value]
            for v in model_instance.sVehicles
            for s in model_instance.sStops
            for w in model_instance.sWarehouses
            for c in model_instance.sCommodities
            if model_instance.vLoadQuantity[v, s, c].value + model_instance.vUnloadQuantity[v, s, c].value +
            model_instance.vQuantityAtArrival[v, s, c].value > 0 and model_instance.vAlpha[v, s, w].value == 1
        ]).to_dictlist(["T_llegada"])

        def convertir_minutos_a_horas(minutos):  # Función para convertir minutos a horas
            horas = (minutos + 1200) // 60
            minutos = (minutos + 1200) % 60
            segundos = int((minutos % 1) * 60)
            if segundos >= 59:  # Redondea posibles decimales
                minutos += 1
                segundos = 0
            if segundos >= 29 and segundos < 31:  # Redondea posibles decimales
                segundos = 30
            horas = horas % 24  # Reiniciar las horas a 0 si superan las 24
            hora_str = "{:02d}:{:02d}:{:02d}".format(int(horas), int(minutos), segundos)
            return hora_str

        for d in data_dep_time:  # Pasamos el df data_dep_time a formato hh:mm:ss
            minutos = d['T_salida']
            hora = convertir_minutos_a_horas(minutos)
            d['T_salida'] = hora

        for d in data_arr_time:  # Pasamos el df data_arr_time a horas a formato hh:mm:ss
            minutos = d['T_llegada']
            hora = convertir_minutos_a_horas(minutos)
            d['T_llegada'] = hora

        warehouses_visited = {(v, s): w for v in model_instance.sVehicles for s in model_instance.sStops for
                              w in model_instance.sWarehouses if model_instance.vAlpha[v, s, w].value == 1}

        trip_durations = {
            (v, s): (model_instance.pTripDuration[warehouses_visited[v, s], warehouses_visited[v, s + 1]].value if
                     (v, s + 1) in warehouses_visited.keys() else 0) for
            v in model_instance.sVehicles for s in model_instance.sStopsButLast
            if (v, s) in warehouses_visited.keys()}

        for v in model_instance.sVehicles:
            trip_durations[v, len(model_instance.sStops) - 1] = 0

        # Definición vble no_opt: número pallets opcionales entregados
        no_opt = sum(model_instance.vUnloadQuantity[v, s, c].value
                     for v in model_instance.sVehicles
                     for s in model_instance.sStops
                     for c in model_instance.sCommodities
                     for w in model_instance.sWarehouses
                     if c[3] == 0 and model_instance.vLoadQuantity[v, s, c].value +
                     model_instance.vUnloadQuantity[v, s, c].value +
                     model_instance.vQuantityAtArrival[v, s, c].value
                     > 0
                     and model_instance.vAlpha[v, s, w].value == 1)

        # Definición vble no_opt_total: número pallets opcionales iniciales
        no_opt_total = self.instance.data["parameters"]["no_opt_total"]

        # Definición vble no_req: número pallets requeridos entregados
        no_req = sum(model_instance.vUnloadQuantity[v, s, c].value
                     for v in model_instance.sVehicles
                     for s in model_instance.sStops
                     for c in model_instance.sCommodities
                     for w in model_instance.sWarehouses
                     if c[3] == 1 and model_instance.vLoadQuantity[v, s, c].value +
                     model_instance.vUnloadQuantity[v, s, c].value +
                     model_instance.vQuantityAtArrival[v, s, c].value
                     > 0
                     and model_instance.vAlpha[v, s, w].value == 1)

        # Definición vble no_req_total: número pallets requeridos iniciales
        no_req_total = self.instance.data["parameters"]["no_req_total"]

        # Definición vbles tiempos:
        t_inicial = data_arr_time[0]["T_llegada"]  # Tiempo (hora) inicial de ventana temporal
        t_final = data_dep_time[-1]["T_salida"]  # Tiempo (hora) final en que se termina el transporte
        t_disponible = self.instance.data["parameters"]["opt_time_limit"] // 60  # Tiempo (horas) disponibles

        # Df con la información detallada de cada vehículo
        data_solution = TupList([[v, s, w, c[0], c[1], c[2], c[3],
                                  model_instance.vLoadQuantity[v, s, c].value,
                                  model_instance.vUnloadQuantity[v, s, c].value]
                                 for v in model_instance.sVehicles
                                 for s in model_instance.sStops
                                 for w in model_instance.sWarehouses
                                 for c in model_instance.sCommodities
                                 if model_instance.vLoadQuantity[v, s, c].value +
                                 model_instance.vUnloadQuantity[v, s, c].value +
                                 model_instance.vQuantityAtArrival[v, s, c].value
                                 > 0
                                 and model_instance.vAlpha[v, s, w].value == 1
                                 ]).to_dictlist(["Vehículo", "Parada", "Localización", "O", "D", "Q", "R", "Carga",
                                                 "Descarga"])

        for i in range(len(data_solution)):
            data_solution[i].update(data_arr_time[i])
            data_solution[i].update(data_dep_time[i])

        df_resumen = pd.DataFrame(data_solution)

        # Df con el esquema de la ruta de cada vehículo
        data_solution_ruta = TupList([[v, s, w]
                                      for v in model_instance.sVehicles
                                      for s in model_instance.sStops
                                      for w in model_instance.sWarehouses
                                      for c in model_instance.sCommodities
                                      if model_instance.vLoadQuantity[v, s, c].value +
                                      model_instance.vUnloadQuantity[v, s, c].value +
                                      model_instance.vQuantityAtArrival[v, s, c].value
                                      > 0
                                      and model_instance.vAlpha[v, s, w].value == 1
                                      ]).to_dictlist(["vehicle", "stop", "warehouse"])
        solucion_ruta = {}

        # Iterate through the data and group stops for each vehicle
        for d in data_solution_ruta:
            vehicle = d['vehicle']
            stop = d['stop']
            warehouse = d['warehouse']

            if vehicle not in solucion_ruta:
                solucion_ruta[vehicle] = {}

            solucion_ruta[vehicle][stop] = warehouse

        df_ruta = pd.DataFrame.from_dict(solucion_ruta, orient='index')
        df_ruta = df_ruta.rename(columns=lambda x: f'{x}')

        # Creación EXCEL
        workbook = Workbook()

        worksheet1 = workbook.active  # Obtener la primera hoja del libro de trabajo
        worksheet1.title = 'Resumen'

        vehiculos = [f'Veh {i}' for i in range(len(df_ruta))]  # Crear una lista con los nombres de los vehículos

        df_ruta.insert(0, "Paradas:", vehiculos)

        worksheet1['A1'] = "Ruta:"

        worksheet1.append([])

        for row in dataframe_to_rows(df_ruta, index=False, header=True):
            worksheet1.append(row)

        worksheet1.append([])

        data_comm = [["Nº pallets", "Iniciales", "Entregados"],
                     ["Obligatorios", no_req_total, no_req],
                     ["Opcionales", no_opt_total, no_opt]]
        columnas = ['Columna 1', 'Columna 2', 'Columna 3']
        df1 = pd.DataFrame(data_comm, columns=columnas)
        for row in dataframe_to_rows(df1, index=False, header=False):
            worksheet1.append(row)

        worksheet1.append([])

        data_time = [[" ", "Disponible", "Utilizado"],
                     ["Tiempo", str(t_disponible) + " h", str(t_utilizado) + " h"]]
        columnas = ['Columna 1', 'Columna 2', 'Columna 3']
        df2 = pd.DataFrame(data_time, columns=columnas)
        for row in dataframe_to_rows(df2, index=False, header=False):
            worksheet1.append(row)

        worksheet1.append([])

        data_time = [[" ", "Inicio", "Fin"],
                     ["Hora", t_inicial, t_final]]
        columnas = ['Columna 1', 'Columna 2', 'Columna 3']
        df2 = pd.DataFrame(data_time, columns=columnas)
        for row in dataframe_to_rows(df2, index=False, header=False):
            worksheet1.append(row)

        # Crear hojas adicionales para los datos de data_solution_resumen
        for vehicle in df_resumen['Vehículo'].unique():
            vehicle_data = df_resumen[df_resumen['Vehículo'] == vehicle]
            worksheet = workbook.create_sheet(title=f'Vehículo {vehicle}')
            for row in dataframe_to_rows(vehicle_data, index=False, header=True):
                worksheet.append(row)

        worksheet1['A3'] = "Paradas: "

        # Alineación texto
        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guardar el libro de trabajo en un archivo Excel
        workbook.save('datos_salida.xlsx')

        data_solution_extra = TupList([[v, s, w, c[0], c[1], c[2], c[3],
                                        model_instance.vQuantityAtArrival[v, s, c].value,
                                        model_instance.vLoadQuantity[v, s, c].value,
                                        model_instance.vUnloadQuantity[v, s, c].value,
                                        model_instance.vArrivalTime[v, s].value,
                                        model_instance.vLoadDuration[v, s].value,
                                        model_instance.vUnloadDuration[v, s].value,
                                        model_instance.vUnloadTime[v, s].value,
                                        model_instance.vDepartureTime[v, s].value,
                                        trip_durations[v, s],
                                        model_instance.vGamma[v, s].value]
                                       for v in model_instance.sVehicles
                                       for s in model_instance.sStops
                                       for w in model_instance.sWarehouses
                                       for c in model_instance.sCommodities
                                       if model_instance.vLoadQuantity[v, s, c].value +
                                       model_instance.vUnloadQuantity[v, s, c].value +
                                       model_instance.vQuantityAtArrival[v, s, c].value
                                       > 0
                                       and model_instance.vAlpha[v, s, w].value == 1
                                       ]).to_dictlist(["vehicle", "stop", "warehouse", "comm_or", "comm_dest",
                                                       "comm_qty", "comm_comp", "qty_arr", "load", "unload", "arr_time",
                                                       "load_dur", "unload_dur", "unload_time", "dep_time", "trip_dur",
                                                       "gamma"])

        # # Crear un objeto ExcelWriter para guardar los datos en un archivo Excel
        # writer = pd.ExcelWriter('solucion_extra.xlsx', engine='xlsxwriter')
        #
        # # Iterar sobre los vehículos y guardar cada uno en una página separada
        # for vehicle in df['vehicle'].unique():
        #     vehicle_data = df[df['vehicle'] == vehicle]
        #     vehicle_data.to_excel(writer, sheet_name=f'Vehicle {vehicle}', index=False)
        #
        # # Guardar y cerrar el archivo Excel
        # writer.save()

        return data_solution_extra

        # Example of solve method:
        #
        # model = create_model()
        #
        # model_instance = model.create_instance(data, report_timing=False)
        # logfile = "./data/logfile.log"
        # # Solve
        # opt = self.set_solver(options)
        # try:
        #     result = opt.solve(
        #         model_instance,
        #         tee=True,
        #         warmstart=False,
        #         logfile=logfile,
        #     )
        # except ApplicationError:
        #     message = "Solver error: a solver license may not be available to solve the model."
        #     raise Exception(message)
        #
        # self.status = get_status(result)
        # model_result = model_instance
        # obj = model_instance.f_obj()
        # print("Status: {} Objective value: {}".format(self.status, obj))
        #
        # # Prepare solution
        # if is_feasible(self.status):
        #     self.totals = self.get_total(model_result, result)
        #     model_solution = self.get_model_solution(model_result)
        #     self.solution = Solution(self.build_solution(model_solution))
        # else:
        #     self.solution = self.get_empty_solution()
        #     self.variables_to_excel(model_result)
        #
        # return dict(status=STATUS_TIME_LIMIT, status_sol=SOLUTION_STATUS_FEASIBLE)
